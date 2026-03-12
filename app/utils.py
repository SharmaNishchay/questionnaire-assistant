import openpyxl
import openpyxl.utils
import csv
from pathlib import Path
from typing import List, Dict
import pypdf


def _extract_rows_from_file(file_path: str) -> List[tuple]:
    """
    Load every row from an Excel (.xlsx / .xls) or CSV file and return them
    as a list of plain tuples.  Raises ValueError with a human-readable
    message if the format is not supported or the file is corrupt.
    """
    suffix = Path(file_path).suffix.lower()

    # .xlsx
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = [tuple(c.value for c in row) for row in ws.iter_rows()]
            wb.close()
            return rows
        except Exception as e:
            # Some .xlsx files are mislabelled old-format .xls — fall through
            if "zip" not in str(e).lower() and "not a zip" not in str(e).lower():
                raise ValueError(f"Cannot read .xlsx file: {e}")
            # fall through to xlrd

    # .xls  (or mislabelled .xlsx that failed above)
    if suffix in (".xls",) or (suffix in (".xlsx",)):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    # xlrd type 1 = text, 2 = number, 3 = date, 0 = empty
                    if cell.ctype == 2 and cell.value == int(cell.value):
                        row.append(int(cell.value))
                    else:
                        row.append(cell.value if cell.ctype != 0 else None)
                rows.append(tuple(row))
            return rows
        except Exception:
            pass  # fall through to CSV last-resort

    # Last resort: treat as CSV (handles files mislabelled as .xlsx but saved as CSV)
    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [tuple(row) for row in reader if any(cell.strip() for cell in row)]
        if rows:
            return rows
    except Exception:
        pass

    raise ValueError(
        f"Cannot read '{Path(file_path).name}'. "
        "Please open it in Excel and save as Excel Workbook (.xlsx), or export as .csv."
    )


def _detect_question_col(all_rows: List[tuple]):
    """
    Return (header_row_idx, question_col, answer_col) — all 0-based.
    Strategy:
      1. Find the first row that has a cell containing 'question' (not '#' / 'number').
      2. Fallback: pick the column with the most long-string values.
    """
    for r_idx, row in enumerate(all_rows):
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        for c_idx, cell_str in enumerate(row_lower):
            if cell_str == "question" or (
                cell_str.startswith("question")
                and "number" not in cell_str
                and "#" not in cell_str
                and "no" not in cell_str
            ):
                answer_col = None
                for ac_idx, astr in enumerate(row_lower):
                    if "answer" in astr or "response" in astr:
                        answer_col = ac_idx
                        break
                return r_idx, c_idx, answer_col

    # Fallback: longest-string column
    col_scores: Dict[int, int] = {}
    for row in all_rows:
        for c_idx, cell in enumerate(row):
            if cell is not None and not isinstance(cell, (int, float)):
                s = str(cell).strip()
                if len(s) > 20:
                    col_scores[c_idx] = col_scores.get(c_idx, 0) + 1
    if col_scores:
        best_col = max(col_scores, key=col_scores.get)
        return 0, best_col, None

    return None, None, None


def parse_questionnaire(file_path: str) -> List[Dict]:
    """
    Parse a questionnaire file (.xlsx, .xls, .csv) and return a list of
    question dicts with keys: number, text, row_data, excel_row.
    Raises ValueError with a user-friendly message on format problems.
    """
    all_rows = _extract_rows_from_file(file_path)
    if not all_rows:
        return []

    header_row_idx, question_col, answer_col = _detect_question_col(all_rows)

    if question_col is None:
        raise ValueError(
            "Could not find a 'Question' column in the file. "
            "Make sure row 3 (or thereabouts) has a header cell containing the word 'Question'."
        )

    data_start = (header_row_idx + 1) if header_row_idx is not None else 0
    questions: List[Dict] = []
    question_num = 1

    for row_idx, row in enumerate(all_rows[data_start:], start=data_start + 1):
        cell_value = row[question_col] if question_col < len(row) else None
        if cell_value is None:
            continue
        question_text = str(cell_value).strip()
        if not question_text:
            continue
        # Skip bare numbers (Q# column mis-identified)
        try:
            float(question_text)
            continue
        except ValueError:
            pass

        row_data = {f"col_{i}": v for i, v in enumerate(row)}
        row_data["_question_col"] = question_col
        row_data["_answer_col"] = answer_col
        row_data["_header_row_idx"] = header_row_idx

        questions.append({
            "number": question_num,
            "text": question_text,
            "row_data": row_data,
            "excel_row": row_idx,
        })
        question_num += 1

    return questions

def export_questionnaire(project_name: str, qa_data: List[Dict]) -> str:
    """Export completed questionnaire to Excel, preserving original structure.

    If the original row data includes column metadata (set by the improved parser),
    we replicate the original workbook's layout and inject answers + citations into
    the answer column (or a new column right after the question column).

    qa_data items: {question_number, question_text, answer_text, citations,
                    confidence_score, original_data}
    """
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in project_name).strip()
    output_path = Path("uploads") / f"{safe_name}_completed.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Completed Questionnaire"

    bold_font = openpyxl.styles.Font(bold=True)
    wrap_align = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    cite_fill = openpyxl.styles.PatternFill("solid", fgColor="EBF5FB")

    # Determine structure from first item's original_data (if available)
    first_od = qa_data[0].get("original_data") if qa_data else {}
    question_col = first_od.get("_question_col") if first_od else None
    answer_col_orig = first_od.get("_answer_col") if first_od else None

    # Decide column indices in output
    # q_out = question col in output, a_out = answer col, c_out = citations, conf_out = confidence
    if question_col is not None:
        q_out = question_col + 1   # openpyxl is 1-based
        # Use the original answer column if it exists, otherwise right after question
        a_out = (answer_col_orig + 1) if answer_col_orig is not None else (q_out + 1)
        c_out = a_out + 1
        conf_out = a_out + 2
    else:
        # Fallback layout: #, Question, Answer, Citations, Confidence
        q_out, a_out, c_out, conf_out = 2, 3, 4, 5

    max_col = max(q_out, a_out, c_out, conf_out)

    # Header row
    ws.cell(row=1, column=1, value="Question #")
    ws.cell(row=1, column=q_out, value="Question")
    ws.cell(row=1, column=a_out, value="Answer")
    ws.cell(row=1, column=c_out, value="Citations")
    ws.cell(row=1, column=conf_out, value="Confidence")
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font

    # Data rows
    for row_num, item in enumerate(qa_data, start=2):
        od = item.get("original_data") or {}
        answer_text = item.get("answer_text") or "Not answered"
        citations = item.get("citations") or []
        confidence = item.get("confidence_score")

        # Reproduce original columns
        for col_key, val in sorted(od.items()):
            if col_key.startswith("col_"):
                col_num = int(col_key.split("_")[1]) + 1
                if col_num <= max_col:
                    ws.cell(row=row_num, column=col_num, value=val)

        # Write question number in column 1
        ws.cell(row=row_num, column=1, value=item.get("question_number", row_num - 1))

        # Ensure question text is always written even if original_data is sparse
        ws.cell(row=row_num, column=q_out, value=item.get("question_text", ""))

        # Answer cell
        ans_cell = ws.cell(row=row_num, column=a_out, value=answer_text)
        ans_cell.alignment = wrap_align

        # Citations cell
        if citations:
            cite_lines = []
            for c in citations:
                src = c.get("source", "")
                snippet = str(c.get("snippet", ""))[:120]
                cite_lines.append(f"[{src}]\n{snippet}{'...' if len(str(c.get('snippet',''))) > 120 else ''}")
            cite_cell = ws.cell(row=row_num, column=c_out, value="\n\n".join(cite_lines))
            cite_cell.alignment = wrap_align
            cite_cell.fill = cite_fill
        else:
            ws.cell(row=row_num, column=c_out, value="—")

        # Confidence cell
        conf_str = f"{confidence * 100:.1f}%" if confidence is not None else "N/A"
        ws.cell(row=row_num, column=conf_out, value=conf_str)

    # Column widths
    col_widths = {q_out: 50, a_out: 60, c_out: 50, conf_out: 12}
    for col_idx in range(1, max_col + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_idx, 20)

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()

    return str(output_path)

def extract_text_from_file(file_path: str) -> str:
    """Extract text from .txt or .pdf file"""
    path = Path(file_path)
    
    if path.suffix.lower() == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    
    elif path.suffix.lower() == ".pdf":
        text = ""
        try:
            reader = pypdf.PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        except Exception as e:
            print(f"Error reading PDF: {e}")
        return text
    
    return ""
