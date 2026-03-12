"""
Script to create sample questionnaire Excel file
Run after installing dependencies
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

def create_sample_questionnaire():
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Security Questionnaire'

    # Add title
    ws['A1'] = 'Vendor Security Assessment Questionnaire'
    ws['A1'].font = Font(size=14, bold=True)

    # Add headers
    ws['A3'] = 'Question #'
    ws['B3'] = 'Question'
    ws['C3'] = 'Answer (To be completed)'

    # Style headers
    for cell in ['A3', 'B3', 'C3']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

    # Questions
    questions = [
        'Does your organization encrypt data at rest? If so, what encryption standard do you use?',
        'How is data encrypted during transmission? Please specify protocols and versions.',
        'What authentication methods do you support for user access?',
        'Do you implement multi-factor authentication (MFA)? Is it required for administrative accounts?',
        'Are you compliant with FERPA regulations? Please describe your compliance measures.',
        'Do you have SOC 2 Type II certification? When was your last audit?',
        'How do you handle COPPA compliance for users under 13 years of age?',
        'Where is your data hosted? Please specify cloud provider and regions.',
        'What is your disaster recovery plan and recovery time objective (RTO)?',
        'How often do you perform security audits and penetration testing?',
        'What is your data retention policy for student records?',
        'Do you provide data portability? In what formats can data be exported?',
    ]

    # Add questions
    for i, question in enumerate(questions, start=4):
        ws[f'A{i}'] = i - 3
        ws[f'B{i}'] = question
        ws[f'C{i}'] = ''

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50

    # Save
    wb.save('sample_data/sample_questionnaire.xlsx')
    print('✓ Sample questionnaire created: sample_data/sample_questionnaire.xlsx')

if __name__ == '__main__':
    create_sample_questionnaire()
